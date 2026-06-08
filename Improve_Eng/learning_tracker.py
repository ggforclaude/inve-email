"""
Improve_Eng/learning_tracker.py
매일 생성된 학습 콘텐츠를 Excel 파일에 자동 기록합니다.

파일 위치: Improve_Eng/learning_log.xlsx  (Git 커밋으로 보관)
시트: 학습이력

컬럼 구성:
  A  일자          | B  Day번호     | C  문법주제(KR)    | D  문법주제(EN)
  E  문법핵심규칙  | F  문법예문    | G  한국인실수패턴  | H  듣기소스
  I  듣기제목      | J  듣기스크립트(EN) | K  듣기핵심어휘  | L  비즈니스표현
  M  비즈니스의미  | N  비즈니스예문    | O  발음포인트     | P  발음규칙
  Q  발음예시      | R  어휘1       | S  어휘2          | T  어휘3
  U  독해전략      | V  독해예시지문    | W  어원단어       | X  어원의미
  Y  어원유래
"""
import os
import logging
import pathlib
from datetime import date

log = logging.getLogger(__name__)

EXCEL_PATH = pathlib.Path(__file__).parent / "learning_log.xlsx"

HEADERS = [
    "일자", "Day", "문법주제(KR)", "문법주제(EN)",
    "문법 핵심규칙", "문법 예문(EN)", "한국인 실수패턴",
    "듣기 소스", "듣기 제목", "듣기 스크립트(EN)", "듣기 핵심어휘",
    "비즈니스 표현", "비즈니스 의미", "비즈니스 예문(EN)",
    "발음 포인트", "발음 규칙", "발음 예시",
    "어휘1 (단어-뜻)", "어휘2 (단어-뜻)", "어휘3 (단어-뜻)",
    "독해 전략", "독해 예시지문",
    "어원 단어", "어원 의미", "어원 유래",
]


def save_learning_log(today: date, day_number: int, learning: dict) -> None:
    """오늘의 학습 콘텐츠를 Excel에 한 행으로 추가."""
    try:
        import openpyxl
        from openpyxl.styles import (Font, PatternFill, Alignment,
                                     Border, Side)
    except ImportError:
        log.error("openpyxl 미설치 — pip install openpyxl")
        return

    grammar  = learning.get("grammar", {})
    listen   = learning.get("listening", {})
    business = learning.get("business", {})
    pron     = learning.get("pronunciation", {})
    vocab    = learning.get("vocabulary", [])
    reading  = learning.get("reading", {})
    etym     = learning.get("etymology_lesson", {})

    # 어휘 3개 텍스트화
    def fmt_vocab(v):
        return f"{v.get('word','')} — {v.get('meaning_kr','')}  |  {v.get('collocation','')}"

    vocab_cols = [fmt_vocab(v) for v in vocab] + ["", "", ""]  # 부족하면 빈칸

    # 문법 예문 첫 번째 (examples 리스트)
    examples = grammar.get("examples", [])
    first_ex = examples[0].get("en", "") if examples else ""

    row = [
        str(today),
        day_number,
        grammar.get("topic_kr", ""),
        grammar.get("topic_en", ""),
        grammar.get("core_rule", ""),
        first_ex,
        grammar.get("common_mistakes", ""),
        listen.get("source", ""),
        listen.get("title", ""),
        listen.get("script_en", ""),
        listen.get("key_vocab", ""),
        business.get("expression", ""),
        business.get("meaning_kr", ""),
        business.get("example_en", ""),
        pron.get("focus", ""),
        pron.get("rule", ""),
        " | ".join(pron.get("examples", [])),
        vocab_cols[0],
        vocab_cols[1],
        vocab_cols[2],
        reading.get("strategy", ""),
        reading.get("example_passage", ""),
        etym.get("word", ""),
        etym.get("meaning_kr", ""),
        etym.get("story_kr", ""),
    ]

    # 파일 열기 or 신규 생성
    if EXCEL_PATH.exists():
        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "학습이력"
        _write_header(ws, HEADERS)

    # 이미 같은 날짜가 있으면 업데이트, 없으면 추가
    date_str = str(today)
    existing_row = None
    for r in ws.iter_rows(min_row=2, max_col=1, values_only=False):
        if r[0].value == date_str:
            existing_row = r[0].row
            break

    if existing_row:
        for col_idx, val in enumerate(row, start=1):
            ws.cell(row=existing_row, column=col_idx, value=val)
        log.info(f"학습 이력 업데이트: {today} (row {existing_row})")
    else:
        ws.append(row)
        new_row = ws.max_row
        _style_data_row(ws, new_row)
        log.info(f"학습 이력 추가: {today} (row {new_row})")

    # 열 너비 자동 조정 (첫 실행 시)
    _set_column_widths(ws)

    wb.save(EXCEL_PATH)
    log.info(f"Excel 저장 완료: {EXCEL_PATH}")


# ── 스타일 헬퍼 ───────────────────────────────────────────────────────────────

def _write_header(ws, headers: list) -> None:
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    header_fill = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
    header_font = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=10)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # 섹션별 색상 매핑
    section_colors = {
        "일자": "1A1A2E", "Day": "1A1A2E",
        "문법주제(KR)": "6D28D9", "문법주제(EN)": "6D28D9",
        "문법 핵심규칙": "6D28D9", "문법 예문(EN)": "6D28D9", "한국인 실수패턴": "6D28D9",
        "듣기 소스": "0284C7", "듣기 제목": "0284C7",
        "듣기 스크립트(EN)": "0284C7", "듣기 핵심어휘": "0284C7",
        "비즈니스 표현": "B45309", "비즈니스 의미": "B45309", "비즈니스 예문(EN)": "B45309",
        "발음 포인트": "BE185D", "발음 규칙": "BE185D", "발음 예시": "BE185D",
        "어휘1 (단어-뜻)": "4338CA", "어휘2 (단어-뜻)": "4338CA", "어휘3 (단어-뜻)": "4338CA",
        "독해 전략": "166534", "독해 예시지문": "166534",
        "어원 단어": "065F46", "어원 의미": "065F46", "어원 유래": "065F46",
    }

    ws.row_dimensions[1].height = 32
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        color = section_colors.get(header, "1A1A2E")
        cell.fill    = PatternFill(start_color=color, end_color=color, fill_type="solid")
        cell.font    = header_font
        cell.alignment = center

    ws.freeze_panes = "C2"  # 일자·Day 고정


def _style_data_row(ws, row_num: int) -> None:
    from openpyxl.styles import Alignment
    wrap = Alignment(vertical="top", wrap_text=True)
    for cell in ws[row_num]:
        cell.alignment = wrap
    ws.row_dimensions[row_num].height = 60


def _set_column_widths(ws) -> None:
    widths = [12, 6, 20, 30, 40, 35, 35,
              20, 30, 50, 30,
              22, 18, 35,
              22, 35, 40,
              35, 35, 35,
              20, 50,
              12, 18, 50]
    for col_idx, width in enumerate(widths, start=1):
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        ws.column_dimensions[col_letter].width = width
